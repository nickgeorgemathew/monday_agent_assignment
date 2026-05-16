"""
prepare_csv.py
Run this ONCE before importing to Monday.com.
Produces: deals_clean.csv and work_orders_clean.csv
"""

import openpyxl
import csv
from datetime import datetime

def fmt_date(val):
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip() if val else ""

def clean_deals(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb["Deal tracker"]
    headers = [ws.cell(1, c).value for c in range(1, 13)]

    skip_values = set(headers)  # skip rows that are header duplicates

    rows = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 13)]
        if row[0] in skip_values or row[0] is None:
            continue

        cleaned = {
            "Deal Name": str(row[0]).strip() if row[0] else "",
            "Owner Code": str(row[1]).strip() if row[1] else "",
            "Client Code": str(row[2]).strip() if row[2] else "",
            "Deal Status": str(row[3]).strip() if row[3] else "",
            "Close Date": fmt_date(row[4]),
            "Closure Probability": str(row[5]).strip() if row[5] else "",
            "Deal Value": str(row[6]).strip() if row[6] else "",
            "Tentative Close Date": fmt_date(row[7]),
            "Deal Stage": str(row[8]).strip() if row[8] else "",
            "Product": str(row[9]).strip() if row[9] else "",
            "Sector": str(row[10]).strip() if row[10] else "",
            "Created Date": fmt_date(row[11]),
        }
        rows.append(cleaned)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    print(f"✅ Deals: {len(rows)} rows written to {output_path}")


def clean_work_orders(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb["work order tracker"]

    # Headers are on row 2
    headers = [ws.cell(2, c).value for c in range(1, 39)]

    rows = []
    for r in range(3, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 39)]
        if row[0] is None or row[0] == headers[0]:
            continue

        cleaned = {
            "Deal Name": str(row[0]).strip() if row[0] else "",
            "Customer Code": str(row[1]).strip() if row[1] else "",
            "Serial #": str(row[2]).strip() if row[2] else "",
            "Nature of Work": str(row[3]).strip() if row[3] else "",
            "Last Recurring Month": str(row[4]).strip() if row[4] else "",
            "Execution Status": str(row[5]).strip() if row[5] else "",
            "Data Delivery Date": fmt_date(row[6]),
            "Date of PO/LOI": fmt_date(row[7]),
            "Document Type": str(row[8]).strip() if row[8] else "",
            "Probable Start Date": fmt_date(row[9]),
            "Probable End Date": fmt_date(row[10]),
            "BD Personnel Code": str(row[11]).strip() if row[11] else "",
            "Sector": str(row[12]).strip() if row[12] else "",
            "Type of Work": str(row[13]).strip() if row[13] else "",
            "Software Platform": str(row[14]).strip() if row[14] else "",
            "Last Invoice Date": fmt_date(row[15]),
            "Latest Invoice No": str(row[16]).strip() if row[16] else "",
            "Amount Excl GST": str(row[17]).strip() if row[17] else "",
            "Amount Incl GST": str(row[18]).strip() if row[18] else "",
            "Billed Excl GST": str(row[19]).strip() if row[19] else "",
            "Billed Incl GST": str(row[20]).strip() if row[20] else "",
            "Collected Amount": str(row[21]).strip() if row[21] else "",
            "To Be Billed Excl GST": str(row[22]).strip() if row[22] else "",
            "To Be Billed Incl GST": str(row[23]).strip() if row[23] else "",
            "Amount Receivable": str(row[24]).strip() if row[24] else "",
            "AR Priority": str(row[25]).strip() if row[25] else "",
            "Quantity by Ops": str(row[26]).strip() if row[26] else "",
            "Quantity as per PO": str(row[27]).strip() if row[27] else "",
            "Quantity Billed": str(row[28]).strip() if row[28] else "",
            "Balance Quantity": str(row[29]).strip() if row[29] else "",
            "Invoice Status": str(row[30]).strip() if row[30] else "",
            "Expected Billing Month": str(row[31]).strip() if row[31] else "",
            "Actual Billing Month": str(row[32]).strip() if row[32] else "",
            "Actual Collection Month": str(row[33]).strip() if row[33] else "",
            "WO Status": str(row[34]).strip() if row[34] else "",
            "Collection Status": str(row[35]).strip() if row[35] else "",
            "Collection Date": fmt_date(row[36]),
            "Billing Status": str(row[37]).strip() if row[37] else "",
        }
        rows.append(cleaned)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    print(f"✅ Work Orders: {len(rows)} rows written to {output_path}")


if __name__ == "__main__":
    clean_deals(
        "../data/Deal_funnel_Data.xlsx",
        "../data/deals_clean.csv"
    )
    clean_work_orders(
        "../data/Work_Order_Tracker_Data.xlsx",
        "../data/work_orders_clean.csv"
    )
